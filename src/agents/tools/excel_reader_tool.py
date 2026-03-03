import io
import logging

from openpyxl import load_workbook
from langchain_core.tools import tool

from src.agents.tools.download_utils import _download_file

logger = logging.getLogger(__name__)

_SAMPLE_ROWS = 5


@tool
def excel_reader_tool(file_url: str) -> str:
    """Download an XLSX from *file_url* and extract sheet metadata with sample rows.

    For each sheet the output includes:
    - Sheet name
    - Column headers (first row)
    - Total row count
    - First 5 data rows as a simple text table

    Returns a structured plain-text string.
    """
    data = _download_file(file_url)

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        if not rows:
            parts.append(f"Sheet: {sheet_name}\n  (empty sheet)\n")
            continue

        headers = rows[0]
        data_rows = rows[1:]
        total_rows = len(data_rows)
        sample = data_rows[:_SAMPLE_ROWS]

        section = [
            f"Sheet: {sheet_name}",
            f"  Columns ({len(headers)}): {', '.join(headers)}",
            f"  Data rows: {total_rows}",
            f"  Sample rows (first {min(_SAMPLE_ROWS, len(sample))}):",
        ]
        for i, r in enumerate(sample, 1):
            section.append(f"    {i}. {' | '.join(r)}")

        parts.append("\n".join(section))

    wb.close()

    result = "\n\n".join(parts)
    logger.info(
        "Extracted metadata for %d sheet(s) from XLSX", len(wb.sheetnames)
    )
    return result
